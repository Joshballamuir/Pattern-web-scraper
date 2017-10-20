#Leadsource email script by Josh Balla-Muir

import smtplib
import openpyxl
from os.path import abspath, expanduser

print('Welcome to Leadsource email automator.')
doc = input("Enter name of spreadsheet located on your Desktop (ensure it is: .xlsx,.xlsm,.xltx, or .xltm): ") 

sv3 = 587
p1 = input("Please enter your email provider (gmail, outlook, aol or yahoo)")
sv = p1.lower()
if sv == 'gmail':
    sv2 = 'smtp.gmail.com'
if sv == 'outlook':
    sv2 = 'smtp.outlook.office365.com'
if sv == 'aol':
    sv2 = 'smtp.aol.com'
if sv == 'yahoo':
    sv2 = 'smtp.mail.yahoo.com'
    sv3 = '465'

print("Please login (remember to allow access for less secure apps)")
un = input("Username:")
pw = input("Password:")
sn = input("Sender name:")
start = int(input("Enter number of start row: "))
end = int(input("Enter number of final row: "))
b = start - 1
loop = 1
while loop == 1:
    
    filepath = abspath(expanduser("~/") + '/Desktop/'+doc)

    wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(filepath)

    b = b+1
    
    ws = wb.active
    name = str((ws['{}{}'.format('b',b)].value)).split(' ', 1)[0]
    emailad = str((ws['{}{}'.format('a',b)].value))
    company = (" ".join([x[0:] for x in str(emailad).split("@")])).split()[1]
    c2 = (" ".join([x[0:] for x in str(company).split(".")])).split()[0]
    c3 = c2+' Ltd'
    c4 = c3.title()
    
    mail = smtplib.SMTP(host=sv2,port=sv3)
    mail.ehlo()
    mail.starttls()

    mail.login(un,pw)
    
    msg = 'Subject: %s\n\n%s' % ('Complaint', "{},""\n""\n""I would like to bring to your attention that I am not satisfied with recent financial dealings I have had with {}.""\n""\n""I wish to make a complaint in regards to a payment. Who is the best person to contact?""\n""\n""Regards,""\n""{}".format(name.title(),c4,sn.title()))
    print('Sending email to the following:')
    print(emailad)

    mail.sendmail(un, emailad, msg) 
    mail.close()                                    
    print(b,'/',end)
    if b == end:
        break

