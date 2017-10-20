#Leadsource email script by Josh Balla-Muir

import smtplib
import openpyxl
from os.path import abspath, expanduser

print('Welcome to Leadsource email automator.')
#doc = input("Enter name of spreadsheet located on your Desktop (ensure it is: .xlsx,.xlsm,.xltx, or .xltm): ") 

sv3 = 587
sv2 = 'smtp.gmail.com'
#p1 = input("Please enter your email provider (gmail, outlook, aol or yahoo)")

 #   sv2 = 'smtp.gmail.com'
#if sv == 'outlook':
#    sv2 = 'smtp.outlook.office365.com'
#if sv == 'aol':
#    sv2 = 'smtp.aol.com'
#if sv == 'yahoo':
#    sv2 = 'smtp.mail.yahoo.com'
#    sv3 = '465'

#print("Please login (remember to allow access for less secure apps)")
#un = input("Username:")
#pw = input("Password:")
start = int(input("Enter number of start row: "))
end = int(input("Enter number of final row: "))
b = start - 1
loop = 1
while loop == 1:
    
    filepath = abspath(expanduser("~/") + '/Desktop/patrick.xlsx')

    wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(filepath)

    b = b+1
    
    ws = wb.active
    emailad = str((ws['{}{}'.format('a',b)].value))
    
    mail = smtplib.SMTP(host=sv2,port=sv3)
    mail.ehlo()
    mail.starttls()

    mail.login('patrick.tullet','ballaandmuir2016')
    
    msg = 'Subject: %s\n\n%s' % ('Contact', "Afternoon""\n""\n""I have been trying to contact you from Ireland today however your office line appears to be down. How can I contact accounts?")
    print('Sending email to the following:')
    print(emailad)

    mail.sendmail('patrick.tullet', emailad, msg) 
    mail.close()                                    
    print(b,'/',end)
    if b == end:
        break
