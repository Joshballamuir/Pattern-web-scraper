#Leadsource search automator by Josh Balla-Muir
import smtplib
import webbrowser
import openpyxl
from os.path import abspath, expanduser
import re
import requests
import urllib.request
from collections import Counter
def startsWith0(element):
    if len(element) == 11:
        return element[0] == '0'
    return False
def startsWith44(element):
    if len(element) == 12:
        return element[0] == '4'
    return False
def startsWith440(element):
    if len(element) == 13:
        return element[0] == '4'
    return False
def startsWith442(element):
    if len(element) == 12:
        return element[1] == '4'
    return False
def startsWith4402(element):
    if len(element) == 13:
        return element[1] == '4'
    return False

loop = 1

print('Welcome to Leadsource search automator.')
doc = input("Enter name of spreadsheet located on your Desktop (ensure it is: .xlsx,.xlsm,.xltx, or .xltm): ") 

while loop == 1:
    start = int(input("Enter number of start row: "))
    end = int(input("Enter number of final row: "))

    filepath = abspath(expanduser("~/") + '/Desktop/'+doc)

    wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(filepath)

    b = start - 1 

    loop = 1
    while loop == 1:
        if b == end:
            un = 'leadsourcedone@gmail.com'
            pw = 'leadsourcedone101'
            msg = 'Subject: %s\n\n%s' % ('DONE', "Check through the leads.")
            mail = smtplib.SMTP(host='smtp.gmail.com',port=587)
            mail.ehlo()
            mail.starttls()
            mail.login(un,pw) 
            mail.sendmail(un, 'joshballamuir@gmail.com', msg) 
            mail.close()
            print('Done!')
            break
        
        b = b+1
        print(b,'/',end)
        ws = wb.active
        company = str((ws['{}{}'.format("a",b)].value)).replace(" ","+")
        company1 = company.replace("&","and")
        address = str((ws['{}{}'.format("d",b)].value)).replace(" ","")
        postcode = str((ws['{}{}'.format("e",b)].value)).replace(" ","")
        
        url = 'https://www.google.co.uk/search?q='+company1+'+'+address+'+'+postcode+'+telephone'
        url2 = 'https://www.google.co.uk/search?q='+company1+'+'+postcode+'+'+'telephone'
        url3 = 'https://www.google.co.uk/search?q='+company1+'+'+'telephone'
        url4 = 'https://www.bing.com/search?q='+company1+'+'+address+'+'+postcode+'+telephone'
        url5 = 'https://www.bing.com/search?q='+company1+'+'+postcode+'+'+'telephone'
        url6 = 'https://www.bing.com/search?q='+company1+'+'+'telephone'

        data = requests.get(url).text+requests.get(url2).text+requests.get(url3).text+requests.get(url4).text+requests.get(url5).text+requests.get(url6).text
        data1 = data.replace("(", "")
        data2 = data1.replace(")","")
        data3 = data2.replace("-", "")
        datap = data3.replace("+","")
        data4 = datap.replace(" ", "")
        
        numbers = list(re.findall('\d{11,13}',data4))
        filt0 = list(filter(startsWith0, numbers))
        filt44 = list(filter(startsWith44, numbers))
        filt440 = list(filter(startsWith440, numbers))
        filt442 = list(filter(startsWith442, filt44))
        filt4402 = list(filter(startsWith4402, filt440))

        filt441 = ("'0".join([x[0:] for x in str(filt442).split("'44")]))
        filt4401 = ("'".join([x[0:] for x in str(filt4402).split("'44")]))

        filt5 = str(filt0)+filt441+filt4401
        filt6 = list(re.findall('\d{11}',filt5))

        data = Counter(filt6)
        list0 = (data.most_common(2))

        list1 = str(list0)
        list2 = list1.replace('(',"")
        list3 = list2.replace(')',"")
        list4 = list3.replace(' ',"")
        list5 = list4.replace("'","")

        if len(list5) < 29:
            ws.cell(row=b, column=11).value = ""
            wb.save(doc)                
        else:
            if list5[13] > list5[27]:
                ws.cell(row=b, column=11).value = list5[1:12]
                wb.save(doc)
            else:
                ws.cell(row=b, column=11).value = ""
                wb.save(doc)
                
    #      TODO:
    #      Efficiency
    #      gui support: number inputs, browse for file
    #      queue next file (+workout spreadsheet limit)
    #      app package
    #      fix large total issue

