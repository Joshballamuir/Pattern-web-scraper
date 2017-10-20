#Leadsource search automator by Josh Balla-Muir
import smtplib
import webbrowser
import openpyxl
from os.path import abspath, expanduser
import re
import requests
import urllib.request
from collections import Counter
def startsWith0(element):
    if len(element) == 11:
        return element[0] == '0'
    return False
def startsWith44(element):
    if len(element) == 12:
        return element[0] == '4'
    return False
def startsWith440(element):
    if len(element) == 13:
        return element[0] == '4'
    return False
def startsWith442(element):
    if len(element) == 12:
        return element[1] == '4'
    return False
def startsWith4402(element):
    if len(element) == 13:
        return element[1] == '4'
    return False

print('Welcome to Leadsource search automator.')
doc = input("Enter name of spreadsheet located on your Desktop (ensure it is: .xlsx,.xlsm,.xltx, or .xltm): ") 

loop = 1
while loop == 1:
    
    start = int(input("Enter number of start row: "))
    end = int(input("Enter number of final row: "))
    b = start - 1 

    filepath = abspath(expanduser("~/") + '/Desktop/'+doc)

    wb = openpyxl.Workbook()
    wb = openpyxl.load_workbook(filepath)

    loop = 1
    while loop == 1:
        if b == end:
            un = 'leadsourcedone@gmail.com'
            pw = 'leadsourcedone101'
            msg = 'Subject: %s\n\n%s' % ('DONE', "Check through the leads.")
            mail = smtplib.SMTP(host='smtp.gmail.com',port=587)
            mail.ehlo()
            mail.starttls()
            mail.login(un,pw) 
            mail.sendmail(un, 'joshballamuir@gmail.com', msg) 
            mail.close()
            print('Done!')
            break
        
        b = b+1
        print(b,'/',end)
        
        ws = wb.active
        
        company = str((ws['{}{}'.format("a",b)].value))
        com1 = str(company.split())
        com2 = com1.replace(',','')
        com3 = com2.replace('[',"")
        com4 = com3.replace(']',"")
        com5 = com4.replace("'","")
        com6 = com5.replace(" ","+")
        company2 = com6.replace("&","and")
        
        address = str((ws['{}{}'.format("e",b)].value))
        ad1 = str(address.split())
        ad2 = ad1.replace(',','')
        ad3 = ad2.replace('[',"")
        ad4 = ad3.replace(']',"")
        ad5 = ad4.replace("'","")
        ad6 = ad5.replace(" ","+")
        address1 = ad6.replace("&","and")
        
        address2 = str((ws['{}{}'.format("f",b)].value))
        ad12 = str(address.split())
        ad22 = ad12.replace(',','')
        ad32 = ad22.replace('[',"")
        ad42 = ad32.replace(']',"")
        ad52 = ad42.replace("'","")
        ad62 = ad52.replace(" ","+")
        address12 = ad62.replace("&","and")
        
        postcode = str((ws['{}{}'.format("i",b)].value))
        po1 = str(address.split())
        po2 = po1.replace(',','')
        po3 = po2.replace('[',"")
        po4 = po3.replace(']',"")
        po5 = po4.replace("'","")
        postcode1 = po5.replace(" ","+")
        
        url1 = 'https://www.google.co.uk/search?q='+company2+'+'+address1+'+'+address12+'+'+postcode1+'+telephone'
        url2 = 'https://www.google.co.uk/search?q='+company2+'+'+postcode1+'+'+'telephone'
        url3 = 'https://www.google.co.uk/search?q='+company2+'+'+'telephone'
        url4 = 'https://www.bing.com/search?q='+company2+'+'+address1+'+'+address12+'+'+postcode1+'+telephone'
        url5 = 'https://www.bing.com/search?q='+company2+'+'+postcode1+'+'+'telephone'
        url6 = 'https://www.bing.com/search?q='+company2+'+'+'telephone'

        data = requests.get(url1).text+requests.get(url2).text+requests.get(url3).text+requests.get(url4).text+requests.get(url5).text+requests.get(url6).text
        data1 = data.replace("(", "")
        data2 = data1.replace(")","")
        data3 = data2.replace("-", "")
        datap = data3.replace("+","")
        data4 = datap.replace(" ", "")
        
        numbers = list(re.findall('\d{11,13}',data4))
        filt0 = list(filter(startsWith0, numbers))
        filt44 = list(filter(startsWith44, numbers))
        filt440 = list(filter(startsWith440, numbers))
        filt442 = list(filter(startsWith442, filt44))
        filt4402 = list(filter(startsWith4402, filt440))
        filt441 = ("'0".join([x[0:] for x in str(filt442).split("'44")]))
        filt4401 = ("'".join([x[0:] for x in str(filt4402).split("'44")]))
        filt5 = str(filt0)+filt441+filt4401
        filt6 = list(re.findall('\d{11}',filt5))

        listnum = Counter(filt6)
        list0 = (listnum.most_common(2))
        list1 = str(list0)
        list2 = list1.replace('(',"")
        list3 = list2.replace(')',"")
        list4 = list3.replace(' ',"")
        list5 = list4.replace("'","")
        list6 = list5.replace(',',' ')
        list7 = list6.replace('[',"")
        list8 = list7.replace(']',"")

        tel = str(list8.split()[0])
        tel2 = list8.split()[2]
        first = int(list8.split()[1])
        second = int(list8.split()[3])
        
        if len(list8) < 27:
            if len(list8) > 11:
                if first > 12:
                    print(com5+' Found')
                    ws.cell(row=b, column=11).value = str(tel)
                    wb.save(filepath)
                else:
                     print(com5+' Not Found')
            else:
                print(com5+' Not Found')                         
        else:
             print(com5+' Found')
             if first > second:
                 ws.cell(row=b, column=11).value = str(tel)
                 wb.save(filepath) 
             if first == second:
                 ws.cell(row=b, column=11).value = str(tel+' '+tel2)
                 wb.save(filepath) 
    #      TODO:
    #      gui support: number inputs, browse for file
    #      queue next file (+workout spreadsheet limit)
    #      app package

