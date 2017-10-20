import smtplib 
import webbrowser #module to use webbrowser features such as url, access to the html scripts
import openpyxl #module to manipulate excel documents
from os.path import abspath, expanduser #module that enables access to file by forging a path
import re #module for a premade search function
import requests #ability to request data directly from a url
import urllib.request #works with the above module
from collections import Counter #module containing mathematical functions. used in this case to count frequency of number occurances

def startsWith0(element):
    if len(element) == 11:
        return element[0] == '0'
    return False #my own definition to filter a number and ensure that it is 11 digits in length and the first character is a '0'
def startsWith44(element):
    if len(element) == 12:
        return element[0] == '4'
    return False #my own definition to filter a number and ensure that it is 12 digits in length and the first character is a '4'
def startsWith441(element):
    if len(element) == 12:
        return element[1] == '4'
    return False #my own definition to filter a number and ensure that it is 12 digits in length and the second character is a '4'
def startsWith440(element):
    if len(element) == 13:
        return element[0] == '4'
    return False #my own definition to filter a number and ensure that it is 13 digits in length and the first character is a '4'
def startsWith4401(element):
    if len(element) == 13:
        return element[1] == '4'
    return False #my own definition to filter a number and ensure that it is 13 digits in length and the second character is a '4'
def startsWith4402(element): 
    if len(element) == 13:
        return element[2] == '0'
    return False #my own definition to filter a number and ensure that it is 13 digits in length and the third character is a '0'

emailad = ''
fname = ''
sname = ''
company = ''
names = []

fsdcom = (fname+sname+'@'+emailad+'.com') #joshballamuir@gmail.com
fsdcoduk = (fname+sname+'@'+emailad+'.co.uk') #joshballamuir@gmail.co.uk
fsdnet = (fname+sname+'@'+emailad+'.net') #joshballamuir@gmail.net
fsdie = (fname+sname+'@'+emailad+'.ie') #joshballamuir@gmail.ie
fsdbiz = (fname+sname+'@'+emailad+'.biz') #joshballamuir@gmail.biz
fsdorg = (fname+sname+'@'+emailad+'.org') #joshballamuir@gmail.org
fsdorgduk = (fname+sname+'@'+emailad+'.org.uk') #joshballamuir@gmail.org.uk        

fdsdcom = (fname+'.'+sname+'@'+emailad+'.com') #josh.ballamuir@gmail.com
fdsdcoduk = (fname+'.'+sname+'@'+emailad+'.co.uk') #josh.ballamuir@gmail.co.uk
fdsdnet = (fname+'.'+sname+'@'+emailad+'.net') #josh.ballamuir@gmail.net
fdsdie = (fname+'.'+sname+'@'+emailad+'.ie') #josh.ballamuir@gmail.ie
fdsdbiz = (fname+'.'+sname+'@'+emailad+'.biz') #josh.ballamuir@gmail.biz
fdsdorg = (fname+'.'+sname+'@'+emailad+'.org') #josh.ballamuir@gmail.org
fdsdorgduk = (fname+'.'+sname+'@'+emailad+'.org.uk') #josh.ballamuir@gmail.org.uk

f0sdcom = (fname[0]+sname+'@'+emailad+'.com') #jballamuir@gmail.com
f0sdcoduk = (fname[0]+sname+'@'+emailad+'.co.uk') #jballamuir@gmail.co.uk
f0sdnet = (fname[0]+sname+'@'+emailad+'.net') #jballamuir@gmail.net
f0sdie = (fname[0]+sname+'@'+emailad+'.ie') #jballamuir@gmail.ie
f0sdbiz = (fname[0]+sname+'@'+emailad+'.biz') #jballamuir@gmail.biz
f0sdorg = (fname[0]+sname+'@'+emailad+'.org') #jballamuir@gmail.org
f0sdorgduk = (fname[0]+sname+'@'+emailad+'.org.uk') #jballamuir@gmail.org.uk

f0dsdcom = (fname[0]+'.'+sname+'@'+emailad+'.com') #j.ballamuir@gmail.com
f0dsdcoduk = (fname[0]+'.'+sname+'@'+emailad+'.co.uk') #j.ballamuir@gmail.co.uk
f0dsdnet = (fname[0]+'.'+sname+'@'+emailad+'.net') #j.ballamuir@gmail.net
f0dsdie = (fname[0]+'.'+sname+'@'+emailad+'.ie') #j.ballamuir@gmail.ie
f0dsdbiz = (fname[0]+'.'+sname+'@'+emailad+'.biz') #j.ballamuir@gmail.biz
f0dsdorg = (fname[0]+'.'+sname+'@'+emailad+'.org') #j.ballamuir@gmail.org
f0dsdorgduk = (fname[0]+'.'+sname+'@'+emailad+'.org.uk') #j.ballamuir@gmail.org.uk

f0dcom = (fname[0]+'@'+emailad+'.com') #j@gmail.com
f0dcoduk = (fname[0]+'@'+emailad+'.co.uk') #j@gmail.co.uk
f0dnet = (fname[0]+'@'+emailad+'.net') #j@gmail.net        
f0die = (fname[0]+'@'+emailad+'.ie') #j@gmail.ie
f0dbiz = (fname[0]+'@'+emailad+'.biz') #j@gmail.biz
f0dorg = (fname[0]+'@'+emailad+'.org') #j@gmail.org
f0dorgduk = (fname[0]+'@'+emailad+'.org.uk') #j@gmail.org.uk

fdcom = (fname+'@'+emailad+'.com') #josh@gmail.com  
fdcoduk = (fname+'@'+emailad+'.co.uk') #josh@gmail.co.uk
fdnet = (fname+'@'+emailad+'.net') #josh@gmail.net
fdie = (fname+'@'+emailad+'.ie') #josh@gmail.ie
fdbiz = (fname+'@'+emailad+'.biz') #josh@gmail.biz
fdorg = (fname+'@'+emailad+'.org') #josh@gmail.org
fdorgduk = (fname+'@'+emailad+'.org.uk') #josh@gmail.org.uk

fb0dcom = (fname+sname[0]+'@'+emailad+'.com') #joshb@gmail.com
fb0dcoduk = (fname+sname[0]+'@'+emailad+'.co.uk') #joshb@gmail.co.uk
fb0dnet = (fname+sname[0]+'@'+emailad+'.net') #joshb@gmail.net
fb0die = (fname+sname[0]+'@'+emailad+'.ie') #joshb@gmail.ie
fb0dbiz = (fname+sname[0]+'@'+emailad+'.biz') #joshb@gmail.biz
fb0dorg = (fname+sname[0]+'@'+emailad+'.org') #joshb@gmail.org
fb0dorgduk = (fname+sname[0]+'@'+emailad+'.org.uk') #joshb@gmail.org.uk

permutations = [fsdcom, fsdcoduk, fsdnet, fsdie,fsdbiz, fsdorg, fsdorgduk, fdsdcom, fdsdcoduk, fdsdnet, fdsdie, fdsdbiz, fdsdorg, fdsdorgduk, f0sdcom, f0dsdcoduk, f0sdnet, f0sdie, f0sdbiz, f0sdorg, f0sdorgduk, f0dsdcom, f0dsdcoduk, f0dsdnet, f0dsdie, f0dsdbiz, f0dsdorg, f0dsdorgduk, f0dcom, f0dcoduk, f0dnet, f0die, f0dbiz, f0dorg, f0dorgduk, fdcom, fdcoduk, fdnet, fdie, fdbiz, fdorg, fdorgduk, fb0dcom, fb0dcoduk, fb0dnet, fb0die, fb0dbiz, fb0dorg ,fb0dorgduk]

print("BEAST MODE LEADSOURCING")

provider = input("Email domain: ")
provider = provider.lower()
po = 587 
#eventually make into a revolving door email system. No login required, cycle between address every few hundred. also explore new addresses.
if provider == 'gmail':
    adrs = 'smtp.gmail.com'
if provider == 'outlook':
    adrs = 'smtp.outlook.office365.com'
if provider == 'aol':
    adrs = 'smtp.aol.com'
if provider == 'yahoo':
    adrs = 'smtp.mail.yahoo.com'
    po = '465'

un = input("Username: ")
pw = input("Password: ")
sn = input("Sender name: ")
mail = smtplib.SMTP(host=adrs,port=po)
mail.ehlo()
mail.starttls()
mail.login(un,pw)

option = 1      
while option == 1:
    print("Find spreadsheets at: https://www.uktradeinfo.com/TradeTools/ImportersDetails/Pages/ImportersSearch.aspx")
    doc = input("Spreadsheet name: ")  
    start = int(input("Start row: "))
    final = int(input("Final row: "))
    row = start - 1
    option = 2
    while option == 2:
        if row == final:
            un1 = 'leadsourcedone@gmail.com'
            pw1 = 'leadsourcedone101'
            msg = 'Subject: %s\n\n%s' % ('DONE', "Check through the leads.")
            mail = smtplib.SMTP(host='smtp.gmail.com',port=587)
            mail.ehlo()
            mail.starttls()
            mail.login(un1,pw1) 
            mail.sendmail(un1, un, msg) 
            print('Done!')
            print('Please wait')
            mail = smtplib.SMTP(host=adrs,port=po)
            mail.ehlo()
            mail.starttls()
            mail.login(un,pw)
            option = 1
        else:
            filepath = abspath(expanduser("~/") + '/Desktop/'+doc) 
            wb = openpyxl.Workbook() 
            wb = openpyxl.load_workbook(filepath) 
            row = row+1 
            ws = wb.active 
            company1 = str((ws['{}{}'.format("a",row)].value))
            address1 = str((ws['{}{}'.format("e",row)].value))
            addressl21 = str((ws['{}{}'.format("f",row)].value))
            postcode1 = str((ws['{}{}'.format("i",row)].value))
            print(row,'/',final)
#######
            com1 = str(company1.split())
            com2 = com1.replace(',','')
            com3 = com2.replace('[',"")
            com4 = com3.replace(']',"")
            com5 = com4.replace(" ","+")
            com6 = com5.replace("&","and")
            company = com6.replace("'","")
                        
            ad1 = str(address1.split())
            ad2 = ad1.replace(',','')
            ad3 = ad2.replace('[',"")
            ad4 = ad3.replace(']',"")
            ad5 = ad4.replace("'","")
            ad6 = ad5.replace(" ","+")
            address = ad6.replace("&","and")
                
            adl21 = str(addressl21.split())
            adl22 = adl21.replace(',','')
            adl23 = adl22.replace('[',"")
            adl24 = adl23.replace(']',"")
            adl25 = adl24.replace("'","")
            adl26 = adl25.replace(" ","+")
            addressl2 = adl26.replace("&","and")
                        
            pc1 = str(postcode1.split())
            pc2 = pc1.replace(',','')
            pc3 = pc2.replace('[',"")
            pc4 = pc3.replace(']',"")
            pc5 = pc4.replace("'","")
            postcode = pc5.replace(" ","+")

            ignore = [(company1.split()[0]),(company1.split()[-1]), 'profiles', 'accountant', 'jobs', 'united', 'kingdom', 'international', 'plc','lp', 'linkedin','ltd','limited','...','/script','/body','director','finance','cfo','chief','financial','officer','controller','accounts','manager','management','fd']

            url1 = 'https://www.google.co.uk/search?q='+com6+'+director+linkedin'
            url2 = 'https://www.google.co.uk/search?q='+com6+'+head+of+finance+linkedin'
            url3 = 'https://www.google.co.uk/search?q='+com6+'+cfo+linkedin'
            url4 = 'https://www.google.co.uk/search?q='+com6+'+chief+financial+officer+linkedin'
            url5 = 'https://www.google.co.uk/search?q='+com6+'+financial+accountant+linkedin'
            url6 = 'https://www.google.co.uk/search?q='+com6+'+finance+director+linkedin'
            url7 = 'https://www.google.co.uk/search?q='+com6+'+financial+controller+linkedin'
            url8 = 'https://www.google.co.uk/search?q='+com6+'+accountant+linkedin'
            url9 = 'https://www.google.co.uk/search?q='+com6+'+head+of+accounts+linkedin'
            url10 = 'https://www.google.co.uk/search?q='+com6+'+accounts+manager+linkedin'
            url11 = 'https://www.google.co.uk/search?q='+com6+'+fd+linkedin'
            url12 = 'https://www.google.co.uk/search?q='+com6+'+management+accountant+linkedin'
            #use COMPANIES HOUSE beta website to increase sample size
            urln1 = 'https://www.google.co.uk/search?q='+company+'+'+address+'+'+addressl2+'+'+postcode+'+telephone'
            urln2 = 'https://www.google.co.uk/search?q='+company+'+'+postcode+'+'+'telephone'
            urln3 = 'https://www.google.co.uk/search?q='+company+'+'+'telephone'
            urln4 = 'https://www.bing.com/search?q='+company+'+'+address+'+'+addressl2+'+'+postcode+'+telephone'
            urln5 = 'https://www.bing.com/search?q='+company+'+'+postcode+'+'+'telephone'
            urln6 = 'https://www.bing.com/search?q='+company+'+'+'telephone'

            data = requests.get(url1).text+requests.get(url2).text+requests.get(url3).text+requests.get(url4).text+requests.get(url5).text+requests.get(url6).text+requests.get(url7).text+requests.get(url8).text+requests.get(url9).text+requests.get(url10).text+requests.get(url11).text+requests.get(url12).text
            mystring = data
            keyword = "<b>LinkedIn</b></a></h3><div"
            before_keyword, keyword, after_keyword = mystring.partition(keyword)
#########
            option = 3
            
            while option == 3:
                
                if (fname.lower() == '/script' and sname.lower() == '/body'):
                    finallist = Counter(names)
                    list0 = (finallist.most_common(2))
                    list01 = str(list0)
                    list2 = list01.replace('(',"")
                    list3 = list2.replace(')',"")
                    list4 = list3.replace("'","")
                    list5 = list4.replace(',',' ')
                    list6 = list5.replace('[',"")
                    list7 = list6.replace(']',"")
                    list8 = list7.split()
                    
                    ead1 = company.replace("-","")
                    ead2 =  ead1.replace("+","")
                    ead3 = ead2.lower()
                    ead4 = ead3.replace("ltd","")
                    emailad = ead4.replace("limited","")
                    
                    fname = str(list8[0])
                    sname = str(list8[1])
                    msg = 'Subject: %s\n\n%s' % ('Unhappy Experience', "{},""\n""\n""I would like to make you aware that recently I have not been happy with financial dealings with your company, {}.""\n""\n""I would like to make a complaint in regards to a payment. Who would be the best person to direct my issues to?""\n""\n""Look forward to hearing from you.""\n""\n""Regards,""\n""{}".format(fname.title(),company1.lower().title(),sn.title()))
                    print(msg)
 #                   mail.sendmail(un, permutations, msg)
                    print(fname,sname,emailad,"-",company1)
                    
                    fname = str(list8[3])
                    sname = str(list8[4])
                    msg = 'Subject: %s\n\n%s' % ('Unhappy Experience', "{},""\n""\n""I would like to make you aware that recently I have not been happy with financial dealings with your company, {}.""\n""\n""I would like to make a complaint in regards to a payment. Who would be the best person to direct my issues to?""\n""\n""Look forward to hearing from you.""\n""\n""Regards,""\n""{}".format(fname.title(),company1.lower().title(),sn.title()))
                    print(msg)
 #                   mail.sendmail(un, permutations, msg)
                    print(fname,sname,emailad,"-",company1)
                    print("Sent!")

                    option = 4
                    while option == 4:
                        data1 = requests.get(urln1).text+requests.get(urln2).text+requests.get(urln3).text+requests.get(urln4).text+requests.get(urln5).text+requests.get(urln6).text
                        dt1 = data1.replace("(", "")
                        dt2 = dt1.replace(")","")
                        dt3 = dt2.replace("-", "")
                        dt4 = dt3.replace("+","")
                        data = dt4.replace(" ", "")
                                    
                        listtosort = list(re.findall('\d{11,13}',data))
                                
                        sw0 = list(filter(startsWith0, listtosort))
                                
                        sw44 = list(filter(startsWith44, listtosort))
                        sw441 = list(filter(startsWith441, sw44))
                                
                        sw440 = list(filter(startsWith440, listtosort))
                        sw4401 = list(filter(startsWith4401, sw440))
                        sw4402 = list(filter(startsWith4402, sw4401))
                                
                        replace441 = ("'0".join([x[0:] for x in str(sw441).split("'44")])) 
                        replace4402 = ("'".join([x[0:] for x in str(sw4402).split("'44")])) 
                        numbers = str(sw0)+replace441+replace4402
                        listsorted = list(re.findall('\d{11}',numbers))

                        finallist = Counter(listsorted)
                        list0 = (finallist.most_common(2))
                        list1 = str(list0)
                        list2 = list1.replace('(',"")
                        list3 = list2.replace(')',"")
                        list4 = list3.replace(' ',"")
                        list5 = list4.replace("'","")
                        list6 = list5.replace(',',' ')
                        list7 = list6.replace('[',"")
                        list8 = list7.replace(']',"")
                        
                        if len(list8) < 27:
                            if len(list8) < 11: 
                                print('Telephone not found')
                                option = 2
                        tel = str(list8.split()[0])
                        tel2 = list8.split()[2]
                        first = int(list8.split()[1])
                        second = int(list8.split()[3])
                                    
                        if len(list8) < 27: 
                            if len(list8) > 11:
                                if first > 12:   
                                    print(str(tel)) 
                                    #print(row=row, column=11).value = str(tel)
 #                                  wb.save(filepath)
                                else:
                                    print('Telephone not found')
                                    option = 2
                            else:
                                print('Telephone not found')
                        else:
                            if first > second:
                                print(str(tel))
                                option = 2
                                #print(row=row, column=11).value = str(tel)
   #                            wb.save(filepath)
                            if first == second:
                                print(str(tel+' '+tel2))
                                option = 2
#                                print(row=row, column=11).value = str(tel+' '+tel2)
#                                wb.save(filepath)
                            
                    option = 2

                else:
                
                    spacer = before_keyword.replace("<b>"," ")
                    spacer2 = spacer.replace("</b>"," ")
                    spacer3 = spacer2.replace("<"," ")
                    spacer4 = spacer3.replace(">"," ")

                    list1 = spacer4.split()
                    fname = list1[-3]
                    sname = list1[-2]
                    newname = (fname.lower()+' '+sname.lower())
                    mystring = after_keyword
                    keyword = "<b>LinkedIn</b></a></h3><div"
                    before_keyword, keyword, after_keyword = mystring.partition(keyword)

                    for i in ignore:
                        if i in (newname):
                            newname = ''
                            
                    names.append(newname.title())
                    if newname == '':
                        names.pop()
            
