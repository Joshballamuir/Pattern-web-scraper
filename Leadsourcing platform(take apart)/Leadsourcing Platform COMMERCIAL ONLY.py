import smtplib #module for sending emails
import webbrowser #module to use webbrowser features such as url, access to the html scripts
import openpyxl #module to manipulate excel documents
from os.path import abspath, expanduser #module that enables access to file by forging a path
import re #module for a premade search function
import requests #ability to request data directly from a url
import urllib.request #works with the above module
from collections import Counter #module containing mathematical functions. used in this case to count frequency of number occurances

def startsWith0(element):
    if len(element) == 11:
        return element[0] == '0'
    return False #my own definition to filter a number and ensure that it is 11 digits in length and the first character is a '0'
def startsWith44(element):
    if len(element) == 12:
        return element[0] == '4'
    return False #my own definition to filter a number and ensure that it is 12 digits in length and the first character is a '4'
def startsWith441(element):
    if len(element) == 12:
        return element[1] == '4'
    return False #my own definition to filter a number and ensure that it is 12 digits in length and the second character is a '4'
def startsWith440(element):
    if len(element) == 13:
        return element[0] == '4'
    return False #my own definition to filter a number and ensure that it is 13 digits in length and the first character is a '4'
def startsWith4401(element):
    if len(element) == 13:
        return element[1] == '4'
    return False #my own definition to filter a number and ensure that it is 13 digits in length and the second character is a '4'
def startsWith4402(element): 
    if len(element) == 13:
        return element[2] == '0'
    return False #my own definition to filter a number and ensure that it is 13 digits in length and the third character is a '0'
permutations = 'none'
custom = 'none'
combine = 'none'
extract = 'none'
emailad = 'none'
fname = 'none'
sname = 'none'
company = 'none'
print("Welcome to Leadsourcing Platform") #welcomes user to platform
provider = 'gmail' #to select their email provider. In this prototype only 4 major providers are officially supported
provider = provider.lower()
po = 587 #by default set to 587 because out of the 4 providers offered, it is onlhy yahoo that will not use it

if provider == 'gmail':
    adrs = 'smtp.gmail.com'
if provider == 'outlook':
    adrs = 'smtp.outlook.office365.com'
if provider == 'aol':
    adrs = 'smtp.aol.com'
if provider == 'yahoo':
    adrs = 'smtp.mail.yahoo.com'
    po = '465'
#matches the selected provider to its corresponding port and address to use in forming emails.

print("Please login with your email account""\n""(Your provider may require you to allow access for less secure apps) ")

un = 'clive.musgrave@gmail.com'
pw = 'samuelsm1th'
sn = 'clive'
#user is asked to enter their login details. the sender name is what will be used at the end of emails if that is what they select
print("Please wait")
mail = smtplib.SMTP(host=adrs,port=po)
mail.ehlo()
mail.starttls()
mail.login(un,pw)
print("You are now logged in.") #starts email system

option = 0
while option == 0:

#starts a loop. imprtant to loop here because it menas the user only needs to login once while this program is open

    option = " "

    if option == "":
        option = input("Press enter for the individual telephone number finder, or enter space for the mass telephone number finder ")
        if option == '':
            print("INDIVIDUAL TELEPHONE NUMBER FINDER")
            company1 = input("Enter the company's name ")
            address1 = input("Enter the first line of the company's address or press enter to skip ")
            addressl21 = input("Enter the second line of the company's address or press enter to skip ")           
            postcode1 = input("Enter the company's postcode or press enter to skip ")
            option = 1 #individual telephone number finder. user must enter details manually if they select this. Result will be returned instantly
            
        if option == ' ':
            print("MASS TELEPHONE NUMBER FINDER")
            option = 2 #mass telephone number finder. user does not enter details as it is extracted from an hmrc spreadsheet
            
    if option == " ":
        option = ""
        if option == '': 
            
            print("EMAIL LIST AUTOMATOR")
            name1 = 'b' #user is asked to note the columns details can be found in for emails. this does not appear for telephone sheets because they follow standard formatting
            name2 = 'c'
            perm = 'yes' #asks if user requires permutations should they not have the email address available
            perm = perm.lower()
            if perm == 'yes':
 #               combine = input("Would you like to use the company names for their domain names? ")
 #               combine = combine.lower()

                #if combine == 'yes':
#                    cold = input("Which column would you like to use? ") #company names and domains are often the same. to make things faster, the user has the option to use one column for both rather than 2 separate ones
 #combines email and company column if user selected
               # if combine == 'no':
                cold = 'd' #user has option to use separate columns for email domain and company name if they prefer. more accurate but more time consuming for user
                colc = 'a'
#extracts email address and company name from separate columns 
                   
            #if perm == 'no':
#                cold = input("Which column can the company's email address be found in? ") #if they know the email addresses they will have made a column in excel for the computer to find them in
#                extract = input("Would you like to use the company names for their domain names? ") #allows them to save time as with the combine option. in this case, the company name is extracted from the email address itself.
#                extract = extract.lower()
#                if extract == 'no':
#                    colc = input("Which column can the company names be found in? ")

            option = 3 #email automator WITH and WITHOUT permutations
            
        if option == ' ':
            print("SINGLE EMAIL SENDER")
            perm = input("Do you require permutations? ")
            perm = perm.lower()
            fname = input("Enter first name ")
            company = input("Enter company name ") #asks for details in order to form an email targetting a single recipient using permutations 
            if perm == 'yes':
                sname = input("Enter second name ")
                emailad = input("Enter mail provider ")
            if perm == 'no': #NEGATES need for last name input!!
                emailad = input("Enter email address ") #asks exact email address to send to if user does not want permutations
            option = 4 #single email sender WITH and WITHOUT permutations

         #LIST OF PREPROGRAMMED PERMUTATIONS:
        #49 at time of writing
            
 
        if option == 4:
            print("Would you like to add any additional (irregular) email addresses? ")
            confirm1 = input("Press enter to send mail or + to add additional addresses ")#gives user to add additional addresses
            confirm1 = confirm1.lower()
            if confirm1 == "":
                if perm == 'yes':
                    mail.sendmail(un, permutations, msg)
                    print("Sent!")
                if perm == 'no':
                    mail.sendmail(un, emailad, msg) 
                    print("Sent!") #sends to default emails                 
            if confirm1 == '+':
                print("Enter up to 5 additional addresses")
                a1 = input("1: ") 
                a2 = input("2: ")
                a3 = input("3: ")
                a4 = input("4: ")
                a5 = input("5: ")
                custom = [a1, a2, a3, a4, a5] #gives user to add up to 5 custom emails, made into a new list
                confirm2 = input("Press enter to add the custom list or enter space to send to custom list only")
                confirm2 = confirm.lower() #user chooses whether to combine custom list with default list or use just custom list
            if confirm2 == '':
                if perm == 'yes':
                    mail.sendmail(un, [permutations,custom], msg) 
                    print("Sent!")
                    option = 0
                if perm == 'no':
                    mail.sendmail(un, [emailad,custom], msg) 
                    print("Sent!")
            if confirm == ' ':
                    mail.sendmail(un, custom, msg) 
                    print("Sent!")
            option = 0 #sends custom emails
            
    if (option == 2 or option == 3):
        if option == 2:
            print("Find spreadsheets at: https://www.uktradeinfo.com/TradeTools/ImportersDetails/Pages/ImportersSearch.aspx")
        doc = 'all contacts 2745.xlsx' #asks the user for the excel spreadsheet filename. in this prototype I have found XLSM ensures optimal speed and compatibility
        start = int(input("Enter number of start row: "))
        final = int(input("Enter number of final row: ")) #user enters rows to start and finish on so they can do as much or little of a task suiting their time needs
        row = start - 2 #in order to loop, row must = row +1 so it must start on row-1.

    while (option == 1 or option == 2 or option == 3):
        if (option == 2 or option == 3):
            if row > final:
                un1 = 'leadsourcedone@gmail.com'
                pw1 = 'leadsourcedone101'
                msg = 'Subject: %s\n\n%s' % ('DONE', "Check through the leads.")
                mail = smtplib.SMTP(host='smtp.gmail.com',port=587)
                mail.ehlo()
                mail.starttls()
                mail.login(un1,pw1) 
                mail.sendmail(un1, un, msg) 
                print('Done!')
                print('Please wait')
                mail = smtplib.SMTP(host=adrs,port=po)
                mail.ehlo()
                mail.starttls()
                mail.login(un,pw)
                option = 0 #sends an email to users address (input at start of program) to notify them when task is complete. Then loops program to selection between email and telephone.
                #this excel loop and email notification system is implemented in multiple different features in the platform. they all access it via a loop for efficiency. this means i only need this system to occur once in my program.
            else:
                filepath = abspath(expanduser("~/") + '/Desktop/'+doc) #in this prototype there is a fixed location to find your spreadsheets #explore writing so as to rename files??
                wb = openpyxl.Workbook() #a shortcut variable is made to call this command later
                wb = openpyxl.load_workbook(filepath) #this works with the above variable to load the excel file into python
                row = row+2 #loops down the rows in excel
                ws = wb.active #activates spreadsheet
                company1 = str((ws['{}{}'.format("a",row)].value))
                address1 = str((ws['{}{}'.format("e",row)].value))
                addressl21 = str((ws['{}{}'.format("f",row)].value))
                postcode1 = str((ws['{}{}'.format("i",row)].value)) #extracts variables required for search from uniformly formatted locations on a spreadsheet
                print(row,'/',final) #outputs current row in console so user can track progress
                
        if option == 3:
            emailad = str((ws['{}{}'.format(cold,row)].value))
            fname = str((ws['{}{}'.format(name1,row)].value))
            sname = str((ws['{}{}'.format(name2,row)].value))
            company = str((ws['{}{}'.format(colc,row)].value))
            if combine == 'yes':
                company = emailad
            if combine == 'no':
                print("c")
            if extract == 'yes':
                c1 = (" ".join([x[0:] for x in str((ws['{}{}'.format(cold,row)].value)).split("@")])).split()[1]
                c2 = (" ".join([x[0:] for x in str(c1).split(".")])).split()[0]
                c3 = c2+' Ltd'
                company = c3.title()
            if extract == 'no':                                                 
                company = str((ws['{}{}'.format(colc,row)].value))
            if perm == 'yes':
                      
            
                fsdcom = (fname+sname+'@'+emailad) #joshballamuir@gmail.com

                fdsdcom = (fname+'.'+sname+'@'+emailad) #josh.ballamuir@gmail.com

                f0sdcom = (fname[0]+sname+'@'+emailad) #jballamuir@gmail.com

                f0dsdcom = (fname[0]+'.'+sname+'@'+emailad) #j.ballamuir@gmail.com

                f0dcom = (fname[0]+'@'+emailad) #j@gmail.com

                fdcom = (fname+'@'+emailad) #josh@gmail.com  

                fb0dcom = (fname+sname[0]+'@'+emailad) #joshb@gmail.com

                sdfcom = (sname+fname+'@'+emailad) #ballamuirjosh@gmail.com

                sdf0com = (sname+fname+'@'+emailad) #ballamuirj@gmail.com

                sdcom = (sname+fname+'@'+emailad) #ballamuir@gmail.com

                sdfdcom = (sname+'.'+fname+'@'+emailad) #ballamuir.josh@gmail.com

                s0dfdcom = (sname[0]+'.'+fname+'@'+emailad) #ballamuir.j@gmail.com

                permutations = [fsdcom,fdsdcom,f0sdcom,f0dsdcom,f0dcom,fdcom,fb0dcom,sdfcom,sdf0com,sdcom,sdfdcom,s0dfdcom]
                #all permutations combined into one variable as a list. This enables them all to be called quickly later on the program

                msg = 'Subject: %s\n\n%s' % ('Unhappy Experience', "{},""\n""\n""I would like to make you aware that recently I have not been happy with financial dealings with your company, {}.""\n""\n""I would like to make a complaint in regards to a payment. Who would be the best person to direct my issues to?""\n""\n""Look forward to hearing from you.""\n""\n""Regards,""\n""{}".format(fname.title(),company.lower().title(),sn.title()))
                #template message that will be sent. blanks are filled in by variables. Email cannot be altered in this prototype because it is the exemplar message and is used as a fixed test variable to ensure all emails format correctly.

                mail.sendmail(un, permutations, msg) #sends emails using permutations
                print(fname,sname,"-",company)
                print("Sent!")

            if perm == 'no':
                mail.sendmail(un, (fname+sname+'@'+emailad) , msg)
                print(fname,sname,"-",company)
                print("Sent!")
           
        if (option == 1 or option == 2):
            com1 = str(company1.split())
            com2 = com1.replace(',','')
            com3 = com2.replace('[',"")
            com4 = com3.replace(']',"")
            com5 = com4.replace("'","")
            com6 = com5.replace(" ","+")
            company = com6.replace("&","and")
                        
            ad1 = str(address1.split())
            ad2 = ad1.replace(',','')
            ad3 = ad2.replace('[',"")
            ad4 = ad3.replace(']',"")
            ad5 = ad4.replace("'","")
            ad6 = ad5.replace(" ","+")
            address = ad6.replace("&","and")
                
            adl21 = str(addressl21.split())
            adl22 = adl21.replace(',','')
            adl23 = adl22.replace('[',"")
            adl24 = adl23.replace(']',"")
            adl25 = adl24.replace("'","")
            adl26 = adl25.replace(" ","+")
            addressl2 = adl26.replace("&","and")
                        
            pc1 = str(postcode1.split())
            pc2 = pc1.replace(',','')
            pc3 = pc2.replace('[',"")
            pc4 = pc3.replace(']',"")
            pc5 = pc4.replace("'","")
            postcode = pc5.replace(" ","+") #takes the data extracted from a spreadsheet and filters them. by removing certain punctuation, this makes the data compatible for a websearch  
                        
            url1 = 'https://www.google.co.uk/search?q='+company+'+'+address+'+'+addressl2+'+'+postcode+'+telephone'
            url2 = 'https://www.google.co.uk/search?q='+company+'+'+postcode+'+'+'telephone'
            url3 = 'https://www.google.co.uk/search?q='+company+'+'+'telephone'
            url4 = 'https://www.bing.com/search?q='+company+'+'+address+'+'+addressl2+'+'+postcode+'+telephone'
            url5 = 'https://www.bing.com/search?q='+company+'+'+postcode+'+'+'telephone'
            url6 = 'https://www.bing.com/search?q='+company+'+'+'telephone' #6 websearch permutations using differenct combinations of the variables to form a search. Two search engines - Google and Bing - are used
            #these permutations and different, independant search engines dramatically increase the sample size taken by the program in an instant. this increases the reliability and accuracy of the information at a speed a human could never compete with

            data1 = requests.get(url1).text+requests.get(url2).text+requests.get(url3).text+requests.get(url4).text+requests.get(url5).text+requests.get(url6).text
            dt1 = data1.replace("(", "")
            dt2 = dt1.replace(")","")
            dt3 = dt2.replace("-", "")
            dt4 = dt3.replace("+","")
            data = dt4.replace(" ", "")
            #this combines the the html data from the first pages of all the web searches into one large string.
            #punctuation is removed because it can camouflage a telephone number - e.g. '+(44) - 7...'
                        
            listtosort = list(re.findall('\d{11,13}',data))
            #the remaining data is sorted into a list consisting only of 11-13 digit numbers
                    
            sw0 = list(filter(startsWith0, listtosort))
            #using my definition, the numbers that are 11 digits long and begin with '0' are placed into a new list
                    
            sw44 = list(filter(startsWith44, listtosort))
            sw441 = list(filter(startsWith441, sw44))
            #using my definition, the numbers that are 12 digits long and begin with '44' are placed into a new list
                    
            sw440 = list(filter(startsWith440, listtosort))
            sw4401 = list(filter(startsWith4401, sw440))
            sw4402 = list(filter(startsWith4402, sw4401))
            #using my definition, the numbers that are 13 digits long and begin with '440' are placed into a new list
                    
            replace441 = ("'0".join([x[0:] for x in str(sw441).split("'44")])) #this converts 12 digit '44' numbers to 11 digit numbers beginning with '0'
            replace4402 = ("'".join([x[0:] for x in str(sw4402).split("'44")])) #this converts 13 digit '440' numbers to 11 digit numbers beginning with '0'

            numbers = str(sw0)+replace441+replace4402
            #all 3 lists are identically formatted: 11 digits and begin with '0'
            listsorted = list(re.findall('\d{11}',numbers))
            #to be safe, the lists are filtered into a single list consisting of 11 digits to ensure no anomalous data passes through

            finallist = Counter(listsorted)
            list0 = (finallist.most_common(2))
            list1 = str(list0)
            list2 = list1.replace('(',"")
            list3 = list2.replace(')',"")
            list4 = list3.replace(' ',"")
            list5 = list4.replace("'","")
            list6 = list5.replace(',',' ')
            list7 = list6.replace('[',"")
            list8 = list7.replace(']',"")
            #out of this list, the two numbers that appear most frequently are selected, with the amount of times they occur displayed after the number itself
            #again, the punctuation is removed
            if len(list8) < 27:
                if len(list8) < 11: #if length is less than 11 then 
                    print('Not found: '+company1.lower().title())
                    option = 2
            tel = str(list8.split()[0])
            tel2 = list8.split()[2]
            first = int(list8.split()[1])
            second = int(list8.split()[3])
            #the comparison data is split into 4 separate variables: the first telephone number, the second telephone number, the frequency of occurance for the first number and the frequency of occurance for the second number
                        
            if len(list8) < 27: 
                if len(list8) > 11:
                    if first > 12: #length of standard comparison return data is 27 characters. therefore if less than this, only one number must have been found
                        #because this is so unlikely, the number is likely anomalous and not correct. In order to be accepted, it must have occured a minimum of 12 times. This means twice per search result on average.
                        print('Found: '+company1.lower().title()) #user is updated in live time when a number is found
                        if option == 1:
                            print(str(tel)) #if option 1 (single number) result is output to console
                            ws.cell(row=row, column=11).value = str(tel)
                            wb.save(filepath) #if using mass finder, result is output to spreadsheet
                            option = 2
                    else:
                        print('Not found: '+company1.lower.title()) #user is updated in live time in the rare occasion number is not found
                        option = 2
                else:
                    print('Not found: '+company1.lower().title())
                    option = 2
            else:
                print('Found: '+company1.lower().title())
                if first > second:
                    if option == 1:
                        print(str(tel)) 
                    if option == 2:
                        ws.cell(row=row, column=11).value = str(tel)
                        wb.save(filepath) #if first number is more frequent, use that number only (second number will never be more frequent)
                if first == second:
                    if option == 1:
                        print(str(tel+' '+tel2))
                    if option == 2:
                        ws.cell(row=row, column=11).value = str(tel+' '+tel2)
                        wb.save(filepath) #if both numbers appear same frequency, use both

            if option == 1:
                option = 0
